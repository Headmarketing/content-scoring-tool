"""
report.py — aggregates per-page scores into a website-level Content Health
Score, builds the priority recommendation list, and renders the dashboard.
"""
import csv
import json
import statistics as st
from collections import Counter, defaultdict
from datetime import datetime
from scoring import CATEGORY_WEIGHTS, ISSUE_LABELS

RECOMMENDATION_MAP = {
    # ── Blog / Article ──
    "thin_content":              ("High",   "Expand thin articles",              "Articles under 800 words rarely rank or fully answer user intent. Expand with sub-sections, examples, and data."),
    "missing_article_schema":    ("High",   "Add Article/BlogPosting schema",     "Add Article or BlogPosting structured data (JSON-LD) so search engines can understand and feature the content."),
    "missing_faq_schema":        ("Medium", "Add FAQPage schema",                 "This page already has FAQ-style content but no matching FAQPage structured data — add it to become eligible for FAQ rich results."),
    "missing_faq":               ("High",   "Add FAQ sections",                   "Add a short FAQ block answering the 3-5 questions readers most often have — improves usefulness and SERP real estate."),
    "no_author":                 ("Medium", "Add author bylines",                 "Show a named author (and ideally credentials) to strengthen E-E-A-T trust signals."),
    "stale_content":             ("Medium", "Refresh outdated content",            "Review and update content older than ~18 months for accuracy and freshness."),
    "few_lists_tables":          ("Low",    "Add visual structure",               "Add bullet lists, numbered steps, or tables to make scanning easier."),
    "long_paragraphs":           ("Medium", "Improve formatting",                  "Break long paragraphs into shorter ones (2-4 sentences) and add more whitespace."),
    "no_cta":                    ("Medium", "Improve CTA placement",              "Add a clear next step for the reader (subscribe, read related post, contact, etc.)."),
    # ── Universal SEO ──
    "no_internal_links":         ("High",   "Improve internal linking",           "Add contextual links to related articles/pages to spread authority and help discovery."),
    "missing_meta_description":  ("High",   "Write meta descriptions",            "Every page should have a unique, compelling meta description (~150-160 characters)."),
    "short_meta_description":    ("Low",    "Lengthen meta descriptions",         "Expand short meta descriptions to closer to 150-160 characters."),
    "missing_h1":                ("High",   "Fix missing H1",                     "Every page needs exactly one clear H1 that matches the page’s primary topic."),
    "multiple_h1":               ("Medium", "Fix heading hierarchy",              "Use a single H1 per page and a logical H2/H3 nesting below it."),
    "missing_images":            ("Medium", "Add supporting visuals",             "Add at least one relevant image, diagram, or screenshot to support the text."),
    "title_missing":             ("High",   "Add a title tag",                    "This page has no meta title. Every page must have a unique, descriptive title tag."),
    "title_too_short":           ("Medium", "Lengthen the title tag",             "Title is under 55 characters. Expand it to 55–65 characters to fully use available SERP space."),
    "title_too_long":            ("Medium", "Shorten the title tag",              "Title exceeds 65 characters and may be truncated in search results. Trim it to 55–65 characters."),
    "missing_alt_text":          ("Medium", "Fix image alt text",                 "Add descriptive alt attributes to content images for accessibility and image search. Decorative images (alt=\"\") are exempt."),
    "missing_breadcrumb":        ("Low",    "Add breadcrumb navigation",          "Add breadcrumb markup/UI to aid navigation and enable breadcrumb rich results."),
    "no_og_tags":                ("Low",    "Add Open Graph tags",                "Add og:title/og:description/og:image so shared links render well on social platforms."),
    # ── Product (e-commerce) ──
    "missing_product_schema":    ("High",   "Add Product schema markup",          "Add Product + Offer JSON-LD to unlock price/availability rich results and improve crawlability."),
    "missing_price_markup":      ("Medium", "Mark up price / offer",              "Use Offer schema or visible, structured price display so users and crawlers can read pricing clearly."),
    "missing_add_to_cart":       ("High",   "Add a clear purchase button",        "Every e-commerce product page needs a prominent Add-to-Cart or Buy Now button visible above the fold."),
    "thin_product_description":  ("Medium", "Expand product description",         "Product descriptions under 150 words limit SEO and may not answer buyer questions. Expand with features, specs, and benefits."),
    # ── Service / Service-based product ──
    "missing_service_schema":    ("Medium", "Add Service schema",                 "Add Service or FinancialProduct JSON-LD markup to help search engines understand the offering."),
    "missing_service_page_schema":("Medium","Add schema for this service offering","This page looks like a service or SaaS product page. Add Service or FinancialProduct JSON-LD markup."),
    "thin_service_description":  ("Medium", "Expand service description",         "Service pages under 300 words rarely rank. Add benefit statements, use-cases, process steps, and FAQs."),
    "service_no_cta":            ("High",   "Add a lead-generation CTA",          "Service pages must have a clear next step: Contact Us, Get a Quote, Book a Demo, See Pricing, etc."),
    # ── Home ──
    "weak_home_cta":             ("High",   "Add a primary CTA to the home page", "The home page must have one clear call-to-action guiding visitors toward the site’s primary goal."),
    "missing_website_schema":    ("Low",    "Add WebSite / Organization schema",  "Add WebSite or Organization JSON-LD on the home page to provide authoritative brand signals."),
    # ── About ──
    "missing_org_schema":        ("Medium", "Add Organization schema",            "Add Organization JSON-LD to the About page to establish brand identity and E-E-A-T signals in search."),
    # ── Contact ──
    "missing_contact_schema":    ("Medium", "Add ContactPage schema",             "Add ContactPage JSON-LD markup so search engines understand this is a contact destination."),
    "no_contact_form":           ("High",   "Add a contact form",                 "Every contact page should include a form. Email and phone alone create friction for users."),
    "missing_contact_info":      ("Medium", "Display phone/email visibly",        "Visible contact information (phone or email) on the contact page builds trust and reduces friction."),
    # ── Community ──
    "thin_community_content":    ("Medium", "Improve community page content",     "Community/forum pages with under 100 words provide little value. Add descriptions, rules, or featured discussions."),
    # ── Generic / Landing ──
    "generic_no_cta":            ("Medium", "Add a call-to-action",               "Even general pages benefit from a clear next step to guide visitor flow."),
}


def build_site_report(pages_with_scores: list[dict]) -> dict:
    """pages_with_scores: list of {**page_data, **score_result}"""
    valid = [p for p in pages_with_scores if not p.get("error")]
    total = len(pages_with_scores)

    if not valid:
        return {"total_articles": total, "error": "No pages could be successfully analyzed."}

    overall_scores = [p["overall_score"] for p in valid]
    cat_avgs = {
        cat: round(st.mean(p["category_scores"][cat] for p in valid), 1)
        for cat in CATEGORY_WEIGHTS
    }

    issue_counter = Counter()
    for p in valid:
        for issue in p["issues"]:
            issue_counter[issue["code"]] += 1

    faq_coverage = round(100 * sum(1 for p in valid if p.get("has_faq")) / len(valid), 1)
    schema_coverage = round(100 * sum(1 for p in valid if p.get("has_article_schema") or p.get("has_faq_schema")) / len(valid), 1)
    internal_link_coverage = round(100 * sum(1 for p in valid if p.get("internal_link_count", 0) >= 2) / len(valid), 1)
    author_coverage = round(100 * sum(1 for p in valid if p.get("author_present")) / len(valid), 1)

    # Priority recommendations: rank issue types by (severity, how many pages affected)
    sev_rank = {"High": 0, "Medium": 1, "Low": 2}
    recommendations = []
    for code, count in issue_counter.items():
        sev, title, desc = RECOMMENDATION_MAP.get(code, ("Low", code, ""))
        recommendations.append({
            "code": code, "severity": sev, "title": title, "description": desc,
            "affected_pages": count, "affected_pct": round(100 * count / len(valid), 1),
        })
    recommendations.sort(key=lambda r: (sev_rank[r["severity"]], -r["affected_pages"]))

    pages_sorted = sorted(valid, key=lambda p: p["overall_score"])
    top_priority_pages = [
        {"url": p["url"], "title": p.get("title") or p["url"], "overall_score": p["overall_score"],
         "issue_count": len(p["issues"]), "word_count": p.get("word_count", 0)}
        for p in pages_sorted[:15]
    ]

    failed = [{"url": p["url"], "error": p.get("error")} for p in pages_with_scores if p.get("error")]

    # Issue log: one row per page; all issues for that page consolidated into
    # a single cell so the sheet shows every problem without duplicate URL rows.
    page_issue_map: dict = defaultdict(lambda: {"title": "", "issues": [], "max_severity": "Low"})
    sev_rank_inner = {"High": 0, "Medium": 1, "Low": 2}
    for p in valid:
        url = p["url"]
        page_issue_map[url]["title"] = p.get("title") or p["url"]
        for issue in p["issues"]:
            page_issue_map[url]["issues"].append(issue["label"])
            if sev_rank_inner.get(issue["severity"], 9) < sev_rank_inner.get(page_issue_map[url]["max_severity"], 9):
                page_issue_map[url]["max_severity"] = issue["severity"]

    issue_log = [
        {
            "url": url,
            "title": data["title"],
            "issue": "; ".join(data["issues"]),
            "severity": data["max_severity"],
        }
        for url, data in page_issue_map.items()
        if data["issues"]
    ]

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total_articles": total,
        "successfully_analyzed": len(valid),
        "failed_pages": failed,
        "overall_health_score": round(st.mean(overall_scores), 1),
        "category_scores": cat_avgs,
        "avg_word_count": round(st.mean(p.get("word_count", 0) for p in valid)),
        "avg_reading_time_min": round(st.mean(p.get("reading_time_min", 0) for p in valid), 1),
        "faq_coverage_pct": faq_coverage,
        "schema_coverage_pct": schema_coverage,
        "internal_link_coverage_pct": internal_link_coverage,
        "author_coverage_pct": author_coverage,
        "issue_distribution": [
            {"code": c, "label": ISSUE_LABELS.get(c, c), "count": n,
             "severity": RECOMMENDATION_MAP.get(c, ("Low",))[0]}
            for c, n in issue_counter.most_common()
        ],
        "recommendations": recommendations,
        "issue_log": issue_log,
        "top_priority_pages": top_priority_pages,
        "content_inventory": [
            {"url": p["url"], "title": p.get("title") or "", "overall_score": p["overall_score"],
             "word_count": p.get("word_count", 0), "reading_time_min": p.get("reading_time_min", 0),
             "category_scores": p["category_scores"], "issue_count": len(p["issues"]),
             "has_faq": p.get("has_faq"), "has_schema": p.get("has_article_schema") or p.get("has_faq_schema"),
             "internal_links": p.get("internal_link_count", 0), "images": p.get("image_count", 0)}
            for p in valid
        ],
    }


def export_xlsx(report: dict, path: str):
    """Excel workbook with 3 tabs: Content Inventory, Recommendations, Issues by Page."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F2937")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Arial", size=10)
    SEV_FILL = {
        "High": PatternFill("solid", fgColor="FBE9E7"),
        "Medium": PatternFill("solid", fgColor="FBEEDB"),
        "Low": PatternFill("solid", fgColor="E4F5F1"),
    }

    def write_sheet(ws, headers, rows, col_widths=None):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for row in rows:
            ws.append(row)
        for r in ws.iter_rows(min_row=2):
            for cell in r:
                cell.font = BODY_FONT
        widths = col_widths or [18] * len(headers)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # --- Tab 1: Summary (mirrors the dashboard's top panels) ---
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.column_dimensions["A"].width = 32
    ws0.column_dimensions["B"].width = 20

    def section(ws, row, title):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        cell.fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        return row + 1

    def kv(ws, row, label, value):
        ws.cell(row=row, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=2, value=value).font = Font(name="Arial", size=10)
        return row + 1

    r = 1
    r = section(ws0, r, f"Content Health Report — {report.get('generated_at', '')}")
    r += 1
    r = kv(ws0, r, "Overall Content Health Score", report.get("overall_health_score"))
    r = kv(ws0, r, "Total Articles", report.get("total_articles"))
    r = kv(ws0, r, "Successfully Analyzed", report.get("successfully_analyzed"))
    r = kv(ws0, r, "Failed to Fetch", len(report.get("failed_pages", [])))
    r += 1

    r = section(ws0, r, "Category Scores (site average, weighted)")
    cat_labels = {"relevance": "Relevance (25%)", "seo": "SEO (20%)", "quality": "Quality (20%)",
                  "usefulness": "Usefulness (25%)", "readability": "Readability (10%)"}
    for cat, label in cat_labels.items():
        r = kv(ws0, r, label, report.get("category_scores", {}).get(cat))
    r += 1

    r = section(ws0, r, "Website-Level Metrics")
    r = kv(ws0, r, "Average Word Count", report.get("avg_word_count"))
    r = kv(ws0, r, "Average Reading Time (min)", report.get("avg_reading_time_min"))
    r += 1

    r = section(ws0, r, "Coverage")
    r = kv(ws0, r, "FAQ Coverage (%)", report.get("faq_coverage_pct"))
    r = kv(ws0, r, "Schema Coverage (%)", report.get("schema_coverage_pct"))
    r = kv(ws0, r, "Internal Link Coverage (%)", report.get("internal_link_coverage_pct"))
    r = kv(ws0, r, "Author Coverage (%)", report.get("author_coverage_pct"))
    r += 1

    if report.get("issue_distribution"):
        r = section(ws0, r, "Issue Distribution (pages affected)")
        for issue in report["issue_distribution"]:
            r = kv(ws0, r, issue["label"], issue["count"])

    # --- Tab 2: Content Inventory ---
    ws1 = wb.create_sheet("Content Inventory")
    inv_headers = ["URL", "Title", "Overall Score", "Relevance", "SEO", "Quality", "Usefulness",
                   "Readability", "Word Count", "Reading Time (min)", "Issue Count",
                   "Has FAQ", "Has Schema", "Internal Links", "Images"]
    inv_rows = [
        [p["url"], p["title"], p["overall_score"], p["category_scores"]["relevance"],
         p["category_scores"]["seo"], p["category_scores"]["quality"], p["category_scores"]["usefulness"],
         p["category_scores"]["readability"], p["word_count"], p["reading_time_min"], p["issue_count"],
         p["has_faq"], p["has_schema"], p["internal_links"], p["images"]]
        for p in report["content_inventory"]
    ]
    write_sheet(ws1, inv_headers, inv_rows, [45, 40, 12, 10, 10, 10, 10, 11, 10, 14, 10, 9, 10, 12, 8])

    # --- Tab 3: Priority Pages (matches dashboard's 'Pages Requiring Priority Updates') ---
    ws1b = wb.create_sheet("Priority Pages")
    pri_headers = ["Rank", "URL", "Title", "Overall Score", "Issue Count", "Word Count"]
    pri_rows = [
        [i + 1, p["url"], p["title"], p["overall_score"], p["issue_count"], p["word_count"]]
        for i, p in enumerate(report.get("top_priority_pages", []))
    ]
    write_sheet(ws1b, pri_headers, pri_rows, [7, 45, 40, 13, 12, 12])
    for idx, p in enumerate(report.get("top_priority_pages", []), start=2):
        v = p["overall_score"]
        fill = SEV_FILL["High"] if v < 50 else (SEV_FILL["Medium"] if v < 75 else SEV_FILL["Low"])
        ws1b.cell(row=idx, column=4).fill = fill

    # --- Tab 4: Recommendations ---
    ws2 = wb.create_sheet("Recommendations")
    rec_headers = ["Priority", "Recommendation", "Description", "Pages Affected", "% of Site"]
    rec_rows = [
        [r["severity"], r["title"], r["description"], r["affected_pages"], r["affected_pct"]]
        for r in report["recommendations"]
    ]
    write_sheet(ws2, rec_headers, rec_rows, [10, 30, 60, 14, 10])
    for i, r in enumerate(report["recommendations"], start=2):
        fill = SEV_FILL.get(r["severity"])
        if fill:
            ws2.cell(row=i, column=1).fill = fill

    # --- Tab 5: Issues by Page ---
    ws3 = wb.create_sheet("Issues by Page")
    issue_headers = ["Severity", "URL", "Title", "Issue"]
    sev_rank = {"High": 0, "Medium": 1, "Low": 2}
    issue_rows_data = sorted(report.get("issue_log", []), key=lambda x: sev_rank.get(x["severity"], 9))
    issue_rows = [[i["severity"], i["url"], i["title"], i["issue"]] for i in issue_rows_data]
    write_sheet(ws3, issue_headers, issue_rows, [10, 45, 40, 45])
    for idx, i in enumerate(issue_rows_data, start=2):
        fill = SEV_FILL.get(i["severity"])
        if fill:
            ws3.cell(row=idx, column=1).fill = fill

    wb.save(path)


def export_csv(report: dict, path: str):
    fields = ["url", "title", "overall_score", "word_count", "reading_time_min",
              "relevance", "seo", "quality", "usefulness", "readability",
              "issue_count", "has_faq", "has_schema", "internal_links", "images"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(fields)
        for row in report["content_inventory"]:
            writer.writerow([
                row["url"], row["title"], row["overall_score"], row["word_count"], row["reading_time_min"],
                row["category_scores"]["relevance"], row["category_scores"]["seo"],
                row["category_scores"]["quality"], row["category_scores"]["usefulness"],
                row["category_scores"]["readability"], row["issue_count"],
                row["has_faq"], row["has_schema"], row["internal_links"], row["images"],
            ])


def render_html(report: dict, out_path: str, site_label: str = "Website"):
    with open(_TEMPLATE_PATH, "r", encoding="utf-8") as f:
        template = f.read()
    html = template.replace("__REPORT_JSON__", json.dumps(report)).replace("__SITE_LABEL__", site_label)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


_TEMPLATE_PATH = __file__.replace("report.py", "dashboard_template.html")
